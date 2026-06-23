"""
Importa os capítulos e páginas especiais a partir da planilha-modelo (.xlsx).

Uso:
    python manage.py import_planilha caminho/para/planilha.xlsx
    python manage.py import_planilha planilha.xlsx --substituir
"""
from django.core.management.base import BaseCommand, CommandError

from content.models import Chapter, SpecialPage


class Command(BaseCommand):
    help = "Importa capítulos e páginas especiais da planilha-modelo (.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str, help="Caminho do arquivo .xlsx")
        parser.add_argument(
            "--substituir",
            action="store_true",
            help="Atualiza capítulos já existentes (por número). Sem isso, pula os que já existem.",
        )

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl não instalado. Rode: pip install openpyxl")

        caminho = opts["arquivo"]
        substituir = opts["substituir"]
        try:
            wb = load_workbook(caminho, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Arquivo não encontrado: {caminho}")

        criados = atualizados = pulados = 0

        # ---- Capítulos ----
        ws = wb["Capítulos"]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            if not row or row[0] in (None, ""):
                continue
            data = dict(zip(headers, row))
            numero = int(data.get("Nº"))
            refs_raw = (data.get("Referências Complementares") or "").strip()
            referencias = "\n".join(
                p.strip() for p in refs_raw.replace(";", "\n").splitlines() if p.strip()
            )
            campos = dict(
                titulo=(data.get("Título") or "").strip(),
                versiculo_texto=(data.get("Versículo-chave (texto)") or "").strip(),
                versiculo_ref=(data.get("Referência") or "").strip(),
                reflexao=(data.get("Reflexão") or "").strip(),
                oracao=(data.get("Oração") or "").strip(),
                aplicacao=(data.get("Aplicação Prática") or "").strip(),
                frase_guardar=(data.get("Frase para Guardar no Coração") or "").strip(),
                referencias=referencias,
                audio_acesso=(data.get("audio_acesso") or "premium").strip(),
                publicado=str(data.get("publicado") or "sim").strip().lower() in ("sim", "true", "1"),
            )
            obj = Chapter.objects.filter(numero=numero).first()
            if obj is None:
                Chapter.objects.create(numero=numero, **campos)
                criados += 1
            elif substituir:
                for k, v in campos.items():
                    setattr(obj, k, v)
                obj.save()
                atualizados += 1
            else:
                pulados += 1

        # ---- Páginas especiais ----
        paginas = 0
        if "Páginas Especiais" in wb.sheetnames:
            ws = wb["Páginas Especiais"]
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c).strip() if c else "" for c in row]
                    continue
                data = dict(zip(headers, row))
                titulo = (data.get("Seção") or "").strip()
                if not titulo:
                    continue
                ordem = int(data.get("Ordem") or 0)
                conteudo = (data.get("Conteúdo") or "").strip()
                _, created = SpecialPage.objects.get_or_create(
                    titulo=titulo, defaults={"conteudo": conteudo, "ordem": ordem}
                )
                if created:
                    paginas += 1

        self.stdout.write(self.style.SUCCESS(
            f"Concluído. Capítulos: {criados} criados, {atualizados} atualizados, "
            f"{pulados} pulados. Páginas especiais: {paginas} criadas."
        ))
